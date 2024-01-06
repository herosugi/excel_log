import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

class ExcelLogger:
    def __init__(self, log_dir="C:\\excel_log", log_filename="work_log.xlsx", headers=None):
        self.log_dir = log_dir
        self.log_filename = log_filename
        
        self.headers = headers if headers is not None else  {
            "Sheet1": ["Datetime", "Header 1", "Header 2", "Header 3", "Header 4", "Header 5"],
            "Sheet2": ["Datetime", "Header 1", "Header 2", "Header 3", "Header 4", "Header 5"]
            }

        self.temp_file_path = os.path.join(self.log_dir, "work_log_temp.xlsx")
                # ワークブックの作成とシートの設定
        self.wb = Workbook()
        self.sheet1 = self.wb.active
        self.sheet1.title = "Sheet1"
        for col_num, header in enumerate(self.headers["Sheet1"], start=1):
            self.sheet1.cell(row=1, column=col_num, value=header)
        
        self.sheet2 = self.wb.create_sheet(title="Sheet2")
        for col_num, header in enumerate(self.headers["Sheet2"], start=1):
            self.sheet2.cell(row=1, column=col_num, value=header)
            

    def log_to_excel(self, data_dict,sheet_name):
        log_file = os.path.join(self.log_dir, self.log_filename)
        temp_file = self.temp_file_path

        # Copy original file to temporary file if it exists
        if os.path.exists(log_file):
            os.replace(log_file, temp_file)

        # Add log to temporary file
        if os.path.exists(temp_file):
            wb = load_workbook(temp_file)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
# Check if the sheet is empty and needs headers
            if ws.max_row == 0 or (ws.max_row == 1 and ws.cell(row=1, column=1).value is None):
                #ws.append(self.headers[sheet_name])
                headers = self.headers[sheet_name]
                for col_num, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col_num).value = header

        else:
            wb = Workbook()
            ws = wb.create_sheet(sheet_name)
            ws.append(self.headers[sheet_name])


        # Automatically add the current datetime in format YYYY-MM-DD HH:MM:SS.sss
        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-4]
        data_dict["Datetime"] = current_datetime

        # Determine the header row based on whether headers were just added
        header_row = 1 if ws.max_row > 1 else 2

        row_data = []
        for header in ws[1]:
            row_data.append(data_dict.get(header.value, ""))

        ws.append(row_data)
        wb.save(temp_file)
        print(f"seve_{sheet_name}")
        # Overwrite original file with temporary file
        if temp_file != log_file:
            os.replace(temp_file, log_file)

    